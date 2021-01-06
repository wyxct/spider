'''
Created on 2019年3月30日

@author: user
'''
import requests
import openpyxl
from lxml import etree
from selenium import webdriver
import time
from selenium.webdriver.support.select import Select
import pprint
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
#https://www.cnblogs.com/zhangxinqi/p/9210211.html 

import win32api,win32con 

def myclick(x,y): 

    win32api.SetCursorPos((x,y)) 

    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN,x,y,0,0) 

    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP,x,y,0,0)  
def get_url_text(url):
    """
    根据url地址， 返回url网页的全部数据
    """
    #返回utf-8的数据
    response = requests.get(url)
    req_text = response.text
    if response.encoding == 'ISO-8859-1':
        ret_text = req_text.encode('ISO-8859-1').decode(response.apparent_encoding)
    else:
        ret_text = req_text      
    return ret_text 

def open_home_brower(url, has_interface = True):
#     chrome_obj = webdriver.Chrome("G://chromedriver_win32//chromedriver.exe") 
    drive_path = "chromedriver.exe"
    if has_interface == False:
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('-no-sanbox')
        brower = webdriver.Chrome(drive_path, chrome_options=chrome_options)
        brower.get(url)
        
        print(brower.page_source)
    else:
        brower = webdriver.Chrome(drive_path)
        brower.get(url)
    time.sleep(2)
    return brower

def select_one_kind_card(brower,select_text):
    select_ele = brower.find_element_by_xpath("//div[@class='select']").click()
    time.sleep(2)
    brower.find_element_by_xpath("//*[@data-text='%s']"%(select_text)).click()
    time.sleep(1)
 
  
def get_current_cars_info(brower): 
    #1 get tite
    text = brower.page_source
    tree = etree.HTML(text)
    table = tree.xpath("//div[@id='tableFixed_tableData']//tr[@class='tr-bg']")
    rows = tree.xpath("//div[@id='tableFixed_tableData']//tbody/tr")
    row = tree.xpath('//*[@id="tableFixed"]/tfoot/tr')
    all_ret = {}
    #for i in range(2,11):
    for i in range(2,11):
        xingcheng_text = table[0].xpath("./td[%s]//p[1]/text()"%(i))[0]
        shijian_text = table[0].xpath("./td[%s]//p[2]/text()"%(i))[0]
        feiyong_text=row[0].xpath("./td[%s]//div/text()"%(i))[0]
        print(xingcheng_text, "  ", shijian_text,'   ',feiyong_text)
        all_ret[xingcheng_text]={}
        all_ret[xingcheng_text]['时间'] = shijian_text
        
        all_ret[xingcheng_text]['发动机机油']='无'
        all_ret[xingcheng_text]['机油滤清器']='无'
        all_ret[xingcheng_text]['变速箱油']='无'
        all_ret[xingcheng_text]['火花塞']='无'
        all_ret[xingcheng_text]['空调滤清器']='无'
        all_ret[xingcheng_text]['空气滤清器']='无'
        all_ret[xingcheng_text]['防冻液']='无'
        all_ret[xingcheng_text]['燃油滤清器']='无'
        all_ret[xingcheng_text]['发动机机油配件费']='无'
        all_ret[xingcheng_text]['发动机机油工时费']='无'
        all_ret[xingcheng_text]['机油滤清器配件费']='无'
        all_ret[xingcheng_text]['机油滤清器工时费']='无'
        all_ret[xingcheng_text]['变速箱油配件费']='无'
        all_ret[xingcheng_text]['变速箱工时费']='无'
        all_ret[xingcheng_text]['火花塞配件费']='无'
        all_ret[xingcheng_text]['火花塞工时费']='无'
        all_ret[xingcheng_text][ '空气滤清器配件费']='无'
        all_ret[xingcheng_text]['空气滤清器工时费']='无'
        all_ret[xingcheng_text]['空凋滤清器配件费']='无'
        all_ret[xingcheng_text]['空调滤清器工时费']='无'
        all_ret[xingcheng_text]['燃油滤清器配件费']='无'
        all_ret[xingcheng_text]['燃油滤清器工时费']='无'
        all_ret[xingcheng_text]['制动液配件费']='无'
        all_ret[xingcheng_text]['制动液工时费']='无'
        all_ret[xingcheng_text]['合计'] = str(feiyong_text)

        lie_num = i - 1
        for j in range(0,(len(rows)-2)):
            all_cl = rows[j].xpath(".//td")
            title = all_cl[0].xpath(".//a/text()")
#             print(title)
            has_or = all_cl[lie_num].xpath(".//p")
            if len(has_or)>= 3:
                peijian_fei = all_cl[lie_num].xpath(".//p[1]/text()")[1]
                gongshi_fei = all_cl[lie_num].xpath(".//p[2]/text()")[0]
                if len(title)==0:
                    continue
                if title[0].strip() in all_ret[xingcheng_text].keys():
                    all_ret[xingcheng_text][title[0].strip()]='有'
                    
                    res = re.search('(\d+\.\d+)',peijian_fei)
                    all_ret[xingcheng_text][title[0].strip()+'配件费']=res.group(1)
                    res = re.search('(\d+\.\d+)',gongshi_fei)
                     
                    all_ret[xingcheng_text][title[0].strip()+'工时费']=res.group(1)
#     pprint.pprint(all_ret)
                                    
    return all_ret    
        
    
def get_cars_detail_from_link(car_link):
    
    #1.判断该车是否是保养
    car_peizhi_all = get_url_text(car_link)
    tree = etree.HTML(car_peizhi_all)
    baoyang_ele = tree.xpath("//*[text()='保养']")
   
    if baoyang_ele[0].tag =='span':
        return False,{}
    #2. 获取保养页面里的所有东西
    baoyang_link = baoyang_ele[0].xpath("./@href")
    print("car_link ",car_link)
    print("baoyang_link ",baoyang_link)
    
#     获取制造商 和 品牌
    title_tmp1 = tree.xpath("//div[@class='subnav-title-name']/a/text()")[0]
    print("zhizhaoby: ",title_tmp1)
    chang_shang = title_tmp1.split("-")[0]
    pingpai = title_tmp1.split("-")[1]
    
    #3. 由于需要JS 渲染， 所以这里启用chrome 获取数据
    
    brower = open_home_brower(baoyang_link[0])
    #跳转到可查找界面
    for_ele = ["//*[text()='2019']","//*[text()='2018']"]
    find_ele = False
    tab_id = 1
    for sig_ele in for_ele:
        try:
            year_ele = brower.find_element_by_xpath(sig_ele)
        except:
            year_ele = False
        if year_ele: 
            year_ele.click()
            time.sleep(2)
            
        else:
            print("has no this cars in this year:", sig_ele)
            continue
        try:
            all_cars_ele = brower.find_element_by_xpath("//div[@class='series-tab-list current']")
        except:
            all_cars_ele = brower.find_element_by_xpath("//div[@id='tab2-2']")
        all_cars_list = all_cars_ele.find_elements_by_tag_name('a')

        can_click = ''
        for sig_car in all_cars_list:
            if sig_car.text.find('暂无') <=0:
                print(sig_car.text)
                can_click = sig_car
                break
        if can_click != '':
            
            can_click.click()
            find_ele = True
            break
    #select all 2018 and 2019 and get all info
    
    if find_ele == False:
        return False, {}
#     get all 2018 cars and 2019 cars text
    all_select_cars = brower.find_elements_by_xpath("//dd[@class='town-btn']/a")
    cars_names = []
    for car_name_ele in all_select_cars:
        car_name = car_name_ele.get_attribute('data-text')
        if car_name.find('2018')==0 or car_name.find('2019')==0:
            cars_names.append(car_name)
    
    ret_cars = []
    for car in cars_names:
        niandai_kuan = car.split(' ')[0]
        chexing = " ".join(car.split(' ')[1:])
        select_one_kind_card(brower,car)
        cars_info = get_current_cars_info(brower)
        curent_cars_info = {}
        curent_cars_info['品牌'] = pingpai
        curent_cars_info['厂商'] = chang_shang
        curent_cars_info['车系'] = pingpai
        curent_cars_info['年代款'] = niandai_kuan
        curent_cars_info['车型'] = chexing
        curent_cars_info['保养信息'] = cars_info
        ret_cars.append(curent_cars_info)
    brower.close()
    return  True,ret_cars
    
               
def get_current_year_cars_link(year):
    url = "https://www.autohome.com.cn/newbrand/0-0-0-%s-0.html"%(year)
    
#     获取改年总共有多少个月有记录
    page_text = get_url_text(url)
    tree = etree.HTML(page_text)
    ret_link = []
    all_care = tree.xpath("//span[text()='配置']")
    print(all_care[0].xpath("../@href"))
    for sig_car in all_care:
        sig_car_link = sig_car.xpath("../@href")[0]
        ret_link.append("https:" + sig_car_link)
    return ret_link
    
def get_all_baoyang():
    int_row = 4
     
    links1 = get_current_year_cars_link('2019')
    links2 = get_current_year_cars_link('2018')
     
    links3 = list(set(links1) | set(links2))
    all_care_info = []        
    for sig_car_link  in  links3:
           
        is_true, sig_car_info = get_cars_detail_from_link(sig_car_link)
        if is_true:
            wb =  load_workbook('汽车之家保养信息(1).xlsx')
            excelsheet = wb.get_sheet_by_name('Sheet1')
            int_row = write_care_info_to_excel(excelsheet, int_row, [sig_car_info])
            wb.save('汽车之家保养信息(1).xlsx')
#             break
        myclick(0, 0)  
        time.sleep(1)
      
    return all_care_info

def write_care_info_to_excel(excelsheet, write_row, all_car_info):
    int_row = write_row
    
    for curent in all_car_info:
        for cars in curent:
            all_xingcheng = cars['保养信息'].keys()
            for xingcheng in all_xingcheng:
                excelsheet.cell(row=int_row, column=column_index_from_string('B')).value = cars['品牌']
                excelsheet.cell(row=int_row, column=column_index_from_string('C')).value = cars['厂商']
                excelsheet.cell(row=int_row, column=column_index_from_string('D')).value = cars['车系']
                excelsheet.cell(row=int_row, column=column_index_from_string('E')).value = cars['车型']
                excelsheet.cell(row=int_row, column=column_index_from_string('F')).value = cars['年代款']
                excelsheet.cell(row=int_row,column = column_index_from_string('G')).value = str(xingcheng)
                excelsheet.cell(row=int_row,column = column_index_from_string('H')).value = str(cars['保养信息'][xingcheng]['时间'])
                excelsheet.cell(row=int_row,column = column_index_from_string('I')).value = cars['保养信息'][xingcheng]['发动机机油']
                excelsheet.cell(row=int_row,column = column_index_from_string('J')).value = cars['保养信息'][xingcheng]['机油滤清器']
                excelsheet.cell(row=int_row,column = column_index_from_string('K')).value = cars['保养信息'][xingcheng]['变速箱油']
                excelsheet.cell(row=int_row,column = column_index_from_string('L')).value = cars['保养信息'][xingcheng]['火花塞']
                excelsheet.cell(row=int_row,column = column_index_from_string('M')).value = cars['保养信息'][xingcheng]['空调滤清器']
                excelsheet.cell(row=int_row,column = column_index_from_string('N')).value = cars['保养信息'][xingcheng]['空气滤清器']
                excelsheet.cell(row=int_row,column = column_index_from_string('P')).value = cars['保养信息'][xingcheng]['燃油滤清器']
                excelsheet.cell(row=int_row,column = column_index_from_string('Q')).value = cars['保养信息'][xingcheng]['发动机机油配件费']
                excelsheet.cell(row=int_row,column = column_index_from_string('R')).value = cars['保养信息'][xingcheng]['发动机机油工时费']
                excelsheet.cell(row=int_row,column = column_index_from_string('S')).value = cars['保养信息'][xingcheng]['机油滤清器配件费']
                excelsheet.cell(row=int_row,column = column_index_from_string('T')).value = cars['保养信息'][xingcheng]['机油滤清器工时费']
                excelsheet.cell(row=int_row,column = column_index_from_string('U')).value = cars['保养信息'][xingcheng]['变速箱油配件费']
                excelsheet.cell(row=int_row,column = column_index_from_string('V')).value = cars['保养信息'][xingcheng]['变速箱工时费']
                excelsheet.cell(row=int_row,column = column_index_from_string('W')).value = cars['保养信息'][xingcheng]['火花塞配件费']
                excelsheet.cell(row=int_row,column = column_index_from_string('X')).value = cars['保养信息'][xingcheng]['火花塞工时费']
                excelsheet.cell(row=int_row,column = column_index_from_string('Y')).value = cars['保养信息'][xingcheng]['空凋滤清器配件费']
                excelsheet.cell(row=int_row,column = column_index_from_string('Z')).value = cars['保养信息'][xingcheng]['空调滤清器工时费']
                excelsheet.cell(row=int_row,column = column_index_from_string('AA')).value = cars['保养信息'][xingcheng]['空气滤清器配件费']
                excelsheet.cell(row=int_row,column = column_index_from_string('AB')).value = cars['保养信息'][xingcheng]['空气滤清器工时费']
                excelsheet.cell(row=int_row,column = column_index_from_string('AC')).value = cars['保养信息'][xingcheng]['燃油滤清器配件费']
                excelsheet.cell(row=int_row,column = column_index_from_string('AD')).value = cars['保养信息'][xingcheng]['燃油滤清器工时费']
                excelsheet.cell(row=int_row,column = column_index_from_string('AE')).value = cars['保养信息'][xingcheng]['燃油滤清器工时费']
                excelsheet.cell(row=int_row,column = column_index_from_string('AF')).value = cars['保养信息'][xingcheng]['制动液工时费']
                excelsheet.cell(row=int_row, column=column_index_from_string('AH')).value = cars['保养信息'][xingcheng]['合计']
                int_row = int_row + 1
    return int_row        
    
    
    
    
    
    
if __name__ == '__main__':
    year = "2019"
#     all_links = get_current_year_cars(year)
#     
#     for sig_link in all_links:
#         get_cars_detail_from_link(sig_link)
        
        
        
#     all = get_cars_detail_from_link('https://car.autohome.com.cn/config/series/620.html#pvareaid=6827543')
#     pprint.pprint(all)


#     brower = open_home_brower('https://car.autohome.com.cn/mtn/series/620#pvareaid=2042316')
#     curent_car = get_current_cars_info(brower)
#     import pprint
#     pprint.pprint(curent_car)
#     import re
#     re_s = re.search('(\d+\.\d+)',"配件费81.00")
#     print(re_s.group(1))

#     write_care_info_to_excel('hehe')
# 
    ALL_CAR = get_all_baoyang()
             
#     write_care_info_to_excel(ALL_CAR)
#     

   
    
       